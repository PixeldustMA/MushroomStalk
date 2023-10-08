from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import sys
import json
from BundleTree import bundleTree
from EditSpreadsheet import edit

# RECIEVE REQUEST
request = ""
for line in sys.stdin:
	request+=line.rstrip()

# PROCESS REQUEST
CharacterRemoval = request.translate({ord(i): None for i in ' []"'})
data = CharacterRemoval.split(",")
dataObject = {
	"TAG": data[9],
	"CHARACTER NAME": data[8].replace("#", " "),
	"CHAPTER NUMBER": data[0],
	"DETAILS": data[1],
	"PATH": str(data[2]),
	"YEAR": data[3],
	"MONTH": data[4],
	"DAY": data[5],
	"TIMEZONE": data[6],
	"UHXTIME": data[10],
	"WEEKDAY": data[11]
}

if data[7] == "Confirm":
	dataObject["DUPLICATE"] = True
else:
	dataObject["DUPLICATE"] = False

# DISPLAY REQUEST
print(json.dumps("The Following Data has been recieved: "))
print(json.dumps(dataObject))

class BundleReader:

	def __init__(self, tag, data, 
				chapterNum, userPath, dateData, 
				timeZoneCode, duplicate, charName,
				UHXTime, weekday):

		''' INITIALISE THE SPREADSHEEET READER'''

		self.TAG = tag
		self.DATA = data
		self.PATH = f'{charName}.xlsx'
		self.CHAPTER_NUM = chapterNum
		self.USERPATH = userPath
		self.WORKBOOK = ""
		self.YEAR = dateData[0]
		self.MONTH = dateData[1]
		self.DAY = dateData[2]
		self.TIMEZONE = timeZoneCode
		self.DUPLICATE = duplicate
		self.CHARACTERNAME = charName
		self.UHXTIME = UHXTime
		self.WEEKDAY = weekday

	def setType(self):

		''' PASS REQUEST TO CORRECT BOOK TYPE'''
		print(json.dumps("Set Type Method Accessed. Finding Correct Tag"))

		match self.TAG:
			case "Bundle":
				self.ReadBundle()
			case "Chrono":
				self.ReadChrono()
			case "Draft":
				self.ReadDraft()

	def getCurrentCell(self, num, step):

		''' GET THE CURRENT CELL '''
		print(json.dumps("Grabbing Current Cell Coordinates"))

		newValue = int(num[1:]) + step
		return num[0] + str(newValue)

	def createHeader(self, active_worksheet):
		headers = {}
		for range in active_worksheet.merged_cells.ranges:
			coords = range.start_cell.coordinate
			details = range.start_cell.value
			headers[coords] = details
		return headers

	def ReadBundle(self):

		''' READ AND ACCESS EDIT METHODS FOR THE DESIRED BUNDLE'''
		print(json.dumps("Bundle Type Accessed"))

		bundle = bundleTree(self.CHARACTERNAME, "Bundle", self.USERPATH).runMatch()

		self.WORKBOOK = load_workbook(filename = bundle)
		active_worksheet = self.WORKBOOK['RECORD']
		Bundle = {
			"LastHeader": "HeaderGoesHere",
			"PreviousChapterCoordinate": "",
			"PreviousChapterDetails": "",
			"CurrentChapterCoordinate": "",
			"CurrentChapterDetails": ""
		}

		for col in active_worksheet.iter_cols(max_col = 1):
			for cell in col:
				if cell.value == int(self.CHAPTER_NUM) - 1:
					Bundle["PreviousChapterCoordinate"] = cell.coordinate

		print(json.dumps(self.DUPLICATE))
		for col in active_worksheet.iter_cols(max_col = 1):
			for cell in col:
				if cell.value == int(self.CHAPTER_NUM):
					if not self.DUPLICATE:
						self.duplicateCheck()
						return ""

		Bundle['CurrentChapterCoordinate'] = self.getCurrentCell(f'{Bundle["PreviousChapterCoordinate"]}', 1)
		Bundle["PreviousChapterDetails"] = active_worksheet[Bundle["PreviousChapterCoordinate"]].value
		Bundle["CurrentChapterDetails"] = self.DATA

		print(json.dumps("BUNDLE READER STATUS..."))
		print(json.dumps(Bundle))

		Editing = edit(self.WORKBOOK, active_worksheet, Bundle["CurrentChapterCoordinate"][1:], self.CHAPTER_NUM)
		Editing.insert()

		Editing.styleSettings()
		AString = f'A{Bundle["CurrentChapterCoordinate"][1:]}' #CHAPTER NUMBER
		Editing.alter(int(self.CHAPTER_NUM), AString) 
		BString = f'B{Bundle["CurrentChapterCoordinate"][1:]}'
		Editing.alter(self.YEAR, BString) 
		CString = f'C{Bundle["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(self.DATA, CString)
		EString = f'E{Bundle["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(int(self.DAY), EString)
		FString = f'F{Bundle["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(int(self.MONTH), FString)
		GString = f'G{Bundle["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(int(self.YEAR), GString)
		HString = f'H{Bundle["CurrentChapterCoordinate"][1:]}' #TIME ZONE CODE
		Editing.alter(self.TIMEZONE, HString)

		self.updateFile(bundle)

	def ReadChrono(self):
		''' READ AND ACCESS EDIT METHODS FOR THE DESIRED CHRONOLOGICAL BOOK'''
		print(json.dumps("Chrono Type Accessed"))

		chrono = bundleTree(self.CHARACTERNAME, "Chrono", self.USERPATH).runMatch()

		self.WORKBOOK = load_workbook(filename = chrono)
		active_worksheet = self.WORKBOOK['SECTION_CODE']

		ChronoObject = {
			"LastHeader": "HeaderGoesHere",
			"PreviousChapterCoordinate": "",
			"PreviousChapterDetails": "",
			"CurrentChapterCoordinate": "",
			"CurrentChapterDetails": ""
		}

		for col in active_worksheet.iter_cols(max_col = 1):
			for cell in col:
				if cell.value == int(self.CHAPTER_NUM) - 1:
					ChronoObject["PreviousChapterCoordinate"] = cell.coordinate

		for col in active_worksheet.iter_cols(max_col = 1):
			for cell in col:
				if cell.value == int(self.CHAPTER_NUM):
					if not self.DUPLICATE:
						self.duplicateCheck()
						return ""

		ChronoObject['CurrentChapterCoordinate'] = self.getCurrentCell(f'{ChronoObject["PreviousChapterCoordinate"]}', 1)
		ChronoObject["PreviousChapterDetails"] = active_worksheet[ChronoObject["PreviousChapterCoordinate"]].value
		ChronoObject["CurrentChapterDetails"] = self.DATA

		print(json.dumps("CHRONO READER STATUS..."))
		print(json.dumps(ChronoObject))

		Editing = edit(self.WORKBOOK, active_worksheet, ChronoObject["CurrentChapterCoordinate"][1:], self.CHAPTER_NUM)
		Editing.insert()

		Editing.styleSettings()
		AString = f'A{ChronoObject["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(int(self.CHAPTER_NUM), AString) 
		BString = f'B{ChronoObject["CurrentChapterCoordinate"][1:]}'
		Editing.alter(self.UHXTIME, BString) 
		CString = f'C{ChronoObject["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(self.YEAR, CString)
		EString = f'D{ChronoObject["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(self.TIMEZONE, EString)
		EString = f'E{ChronoObject["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(int(self.DATA), EString)

		GString = f'G{ChronoObject["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(int(self.WEEKDAY), GString)
		HString = f'H{ChronoObject["CurrentChapterCoordinate"][1:]}'
		Editing.alter(self.DAY, HString)
		HString = f'I{ChronoObject["CurrentChapterCoordinate"][1:]}' 
		Editing.alter(self.MONTH, HString)
		HString = f'J{ChronoObject["CurrentChapterCoordinate"][1:]}'
		Editing.alter(self.YEAR, HString)
		HString = f'K{ChronoObject["CurrentChapterCoordinate"][1:]}'
		Editing.alter(self.TIMEZONE, HString)

	def ReadDraft(self):
		print("Reading Draft")
	
	def findLastHeader(self):
		# cell = sheet.cell(row=row, column=col)
		# if isinstance(cell, MergedCell):
		# 	for merged_range in sheet.merged_cells.ranges:
        #     	if cell.coordinate in merged_range:
        #         	# return the left top cell
        #         	cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
        #         	break
	    # return cell.value
		Headers = sheet_ranges.merged_cells.ranges
		print(Headers)
	def updateFile(self, path):
		self.WORKBOOK.save(path)
	def duplicateCheck(self):
		''' CHECK FOR DUPLICATES '''
		print(json.dumps("Checking For Duplicates"))
		print(json.dumps("DUPLICATE CHAPTER FOUND"))


BundleReader(
	dataObject["TAG"], 
	dataObject["DETAILS"],
	dataObject["CHAPTER NUMBER"], 
	dataObject["PATH"],
	[dataObject["YEAR"], dataObject["MONTH"], dataObject["DAY"]],
	dataObject["TIMEZONE"], 
	dataObject["DUPLICATE"], 
	dataObject["CHARACTER NAME"],
	dataObject["UHXTIME"],
	dataObject["WEEKDAY"]).ReadBundle()