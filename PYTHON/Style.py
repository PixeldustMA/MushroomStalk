from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
import json
black = '000000'
white = 'FFFFFF'

# == // PINK CELLS // == #
PinkStyle = NamedStyle(name = "PinkStyle")
PinkStyle.font = Font(size = 16, name = 'Centralia', color = white)
PinkStyle.fill = PatternFill(start_color='EA157A', end_color='EA157A', fill_type = "solid")
PinkStyle.alignment = Alignment(horizontal="center", vertical="center")
PinkStyle.number_format = "0"

PinkLightStyle = NamedStyle(name = "PinkLightStyle")
PinkLightStyle.font = Font(size = 11, name = 'Centralia', color = white)
PinkLightStyle.fill = PatternFill(start_color='FFACEB', end_color='FFACEB', fill_type = "solid")
PinkLightStyle.alignment = Alignment(horizontal="center", vertical="center")
PinkLightStyle.number_format = "0"

PinkMidStyle = NamedStyle(name = "PinkMidStyle")
PinkMidStyle.font = Font(size = 11, name = 'Centralia', color = black)
PinkMidStyle.fill = PatternFill(start_color='FFD5F4', end_color='FFD5F4', fill_type = "solid")
PinkMidStyle.alignment = Alignment(horizontal="center", vertical="center")

# == // BLUE CELLS // == #
BlueStyle = NamedStyle(name = "BlueStyle")
BlueStyle.font = Font(size = 16, name = 'Centralia', color= white)
BlueStyle.fill = PatternFill(start_color = '000099', end_color='000099', fill_type = "solid")
BlueStyle.alignment = Alignment(horizontal="center", vertical="center")
BlueStyle.number_format = "0"

BlueLightStyle = NamedStyle(name = "BlueLightStyle")
BlueLightStyle.font = Font(size = 11, name = 'Centralia', color= white)
BlueLightStyle.fill = PatternFill(start_color='6F6FFF', end_color='6F6FFF', fill_type = "solid")
BlueLightStyle.alignment = Alignment(horizontal="center", vertical="center")
BlueLightStyle.number_format = "0"

BlueMidStyle = NamedStyle(name = "BlueMidStyle")
BlueMidStyle.font = Font(size = 11, name = 'Centralia', color = black)
BlueMidStyle.fill = PatternFill(start_color='B7B7FF', end_color='B7B7FF', fill_type = "solid")
BlueMidStyle.alignment = Alignment(horizontal="center", vertical="center")

# == // ORANGE CELLS // == #
OrangeStyle = NamedStyle(name = "OrangeStyle")
OrangeStyle.font = Font(size = 16, name = 'Centralia', color = white)
OrangeStyle.fill = PatternFill(start_color = 'FF9933', end_color='FF9933', fill_type = "solid")
OrangeStyle.alignment = Alignment(horizontal="center", vertical="center")
OrangeStyle.number_format = "0"

OrangeLightStyle = NamedStyle(name = "OrangeLightStyle")
OrangeLightStyle.font = Font(size = 11, name = 'Centralia', color= white)
OrangeLightStyle.fill = PatternFill(start_color='FFD6AC', end_color='FFD6AC', fill_type = "solid")
OrangeLightStyle.alignment = Alignment(horizontal="center", vertical="center")
OrangeLightStyle.number_format = "0"

OrangeMidStyle = NamedStyle(name = "OrangeMidStyle")
OrangeMidStyle.font = Font(size = 11, name = 'Centralia', color = black)
OrangeMidStyle.fill = PatternFill(start_color='FFEAD5', end_color='FFEAD5', fill_type = "solid")
OrangeMidStyle.alignment = Alignment(horizontal="center", vertical="center")

# == // GREEN CELLS // == #
GreenStyle = NamedStyle(name = "GreenStyle")
GreenStyle.font = Font(size = 16, name = 'Centralia', color = black)
GreenStyle.fill = PatternFill(start_color='00FF99', end_color='00FF99', fill_type = "solid")
GreenStyle.alignment = Alignment(horizontal="center", vertical="center")
GreenStyle.number_format = "0"

GreenLightStyle = NamedStyle(name = "GreenLightStyle")
GreenLightStyle.font = Font(size = 11, name = 'Centralia', color = white)
GreenLightStyle.fill = PatternFill(start_color='99FFD7', end_color='99FFD7', fill_type = "solid")
GreenLightStyle.alignment = Alignment(horizontal="center", vertical="center")
GreenLightStyle.number_format = "0"

GreenMidStyle = NamedStyle(name = "GreenMidStyle")
GreenMidStyle.font = Font(size = 11, name = 'Centralia', color = '000000')
GreenMidStyle.fill = PatternFill(start_color='CCFFEB', end_color='CCFFEB', fill_type = "solid")
GreenMidStyle.alignment = Alignment(horizontal="center", vertical="center")

# == // PURPLE CELLS // == #
PurpleStyle = NamedStyle(name = "PurpleStyle")
PurpleStyle.font = Font(size = 16, name = 'Centralia', color = black)
PurpleStyle.fill = PatternFill(start_color='CC00FF', end_color='CC00FF', fill_type = "solid")
PurpleStyle.alignment = Alignment(horizontal="center", vertical="center")
PurpleStyle.number_format = "0"

PurpleLightStyle = NamedStyle(name = "PurpleLightStyle")
PurpleLightStyle.font = Font(size = 11, name = 'Centralia', color= white)
PurpleLightStyle.fill = PatternFill(start_color='EB99FF', end_color='EB99FF', fill_type = "solid")
PurpleLightStyle.alignment = Alignment(horizontal="center", vertical="center")
PurpleLightStyle.number_format = "0"

PurpleMidStyle = NamedStyle(name = "PurpleMidStyle")
PurpleMidStyle.font = Font(size = 11, name = 'Centralia', color= white)
PurpleMidStyle.fill = PatternFill(start_color='F4CCFF', end_color='F4CCFF', fill_type = "solid")
PurpleMidStyle.alignment = Alignment(horizontal="center", vertical="center")

# == // PALE BLUE CELLS // == #
PaleBlueStyle = NamedStyle(name = "PaleBlueStyle")
PaleBlueStyle.font = Font(size = 16, name = 'Centralia', color = black)
PaleBlueStyle.fill = PatternFill(start_color='00ADDC', end_color='00ADDC', fill_type = "solid")
PaleBlueStyle.alignment = Alignment(horizontal="center", vertical="center")
PaleBlueStyle.number_format = "0"

PaleBlueLightStyle = NamedStyle(name = "PaleBlueLightStyle")
PaleBlueLightStyle.font = Font(size = 11, name = 'Centralia', color= white)
PaleBlueLightStyle.fill = PatternFill(start_color='8AE4FF', end_color='8AE4FF', fill_type = "solid")
PaleBlueLightStyle.alignment = Alignment(horizontal="center", vertical="center")
PaleBlueLightStyle.number_format = "0"

PaleBlueMidStyle = NamedStyle(name = "PaleBlueMidStyle")
PaleBlueMidStyle.font = Font(size = 11, name = 'Centralia', color= white)
PaleBlueMidStyle.fill = PatternFill(start_color='C4F1FF', end_color='C4F1FF', fill_type = "solid")
PaleBlueMidStyle.alignment = Alignment(horizontal="center", vertical="center")

# == // PALE PINK CELLS // == #
PalePinkStyle = NamedStyle(name = "PalePinkStyle")
PalePinkStyle.font = Font(size = 16, name = 'Centralia', color = black)
PalePinkStyle.fill = PatternFill(start_color='FF33CC', end_color='FF33CC', fill_type = "solid")
PalePinkStyle.alignment = Alignment(horizontal="center", vertical="center")
PalePinkStyle.number_format = "0"

PalePinkLightStyle = NamedStyle(name = "PalePinkLightStyle")
PalePinkLightStyle.font = Font(size = 11, name = 'Centralia', color= white)
PalePinkLightStyle.fill = PatternFill(start_color='F7A2CA', end_color='F7A2CA', fill_type = "solid")
PalePinkLightStyle.alignment = Alignment(horizontal="center", vertical="center")
PalePinkLightStyle.number_format = "0"

PalePinkMidStyle = NamedStyle(name = "PalePinkMidStyle")
PalePinkMidStyle.font = Font(size = 11, name = 'Centralia', color = white)
PalePinkMidStyle.fill = PatternFill(start_color='FBD0E4', end_color='FBD0E4', fill_type = "solid")
PalePinkMidStyle.alignment = Alignment(horizontal="center", vertical="center")

# == // GREY CELLS // == #
GreyStyle = NamedStyle(name = "GreyStyle")
GreyStyle.font = Font(size = 16, name = 'Centralia', color= white)
GreyStyle.fill = PatternFill(start_color='808080', end_color='808080', fill_type = "solid")
GreyStyle.alignment = Alignment(horizontal="center", vertical="center")

class Style:

	def __init__(self, workbook, num, styleName, cell, active):
		self.WORKBOOK = workbook
		self.NUM = num
		self.STYLE_NAME = styleName
		self.CELL = cell,
		self.ACTIVE = active

	def styles(self):
		''' PREPARE WORKBOOK '''

		if not 'PinkStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PinkStyle)
		if not 'BlueStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(BlueStyle)
		if not 'GreenStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(GreenStyle)
		if not 'PurpleStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PurpleStyle)
		if not 'PaleBlueStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PaleBlueStyle)
		if not 'PalePinkStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PalePinkStyle)
		if not 'OrangeStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(OrangeStyle)

		if not 'PinkLightStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PinkLightStyle)
		if not 'BlueLightStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(BlueLightStyle)
		if not 'GreenLightStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(GreenLightStyle)
		if not 'PurpleLightStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PurpleLightStyle)
		if not 'PaleBlueLightStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PaleBlueLightStyle)
		if not 'PalePinkLightStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PalePinkLightStyle)
		if not 'OrangeLightStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(OrangeLightStyle)

		if not 'PinkMidStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PinkMidStyle)
		if not 'BlueMidStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(BlueMidStyle)
		if not 'GreenMidStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(GreenMidStyle)
		if not 'PurpleMidStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PurpleMidStyle)
		if not 'PaleBlueMidStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PaleBlueMidStyle)
		if not 'PalePinkMidStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(PalePinkMidStyle)
		if not 'OrangeMidStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(OrangeMidStyle)

		if not 'GreyStyle' in self.WORKBOOK.named_styles:
			self.WORKBOOK.add_named_style(GreyStyle)

	def setStyle(self):
		''' APPLY STYLE TO A CELL'''
		print(json.dumps("Setting Style"))

		val = self.CELL[0]
		self.ACTIVE[f'A{val}'].style = self.STYLE_NAME
		self.ACTIVE[f'A{val}'].number_format = '0'
		self.additionalStyles()

	def additionalStyles(self):
		''' AUTOMATICALLY SET STYLES'''
		val = self.CELL[0]
		self.ACTIVE[f'D{val}'].style = 'GreyStyle'
		match self.STYLE_NAME:
			case "PinkStyle":
				self.ACTIVE[f'B{val}'].style = 'PinkLightStyle'
				self.ACTIVE[f'C{val}'].style = 'OrangeMidStyle'
				self.ACTIVE[f'E{val}'].style = 'PinkMidStyle'
				self.ACTIVE[f'F{val}'].style = 'PinkMidStyle'
				self.ACTIVE[f'F{val}'].number_format = '0'
				self.ACTIVE[f'G{val}'].style = 'PinkMidStyle'
				self.ACTIVE[f'G{val}'].number_format = '0'
				self.ACTIVE[f'H{val}'].style = 'OrangeMidStyle'
				self.ACTIVE[f'H{val}'].number_format = '0'
			case "BlueStyle":
				self.ACTIVE[f'B{val}'].style = 'BlueLightStyle'
				self.ACTIVE[f'C{val}'].style = 'PinkMidStyle'
				self.ACTIVE[f'E{val}'].style = 'BlueMidStyle'
				self.ACTIVE[f'F{val}'].style = 'BlueMidStyle'
				self.ACTIVE[f'F{val}'].number_format = '0'
				self.ACTIVE[f'G{val}'].style = 'BlueMidStyle'
				self.ACTIVE[f'G{val}'].number_format = '0'
				self.ACTIVE[f'H{val}'].style = 'PinkMidStyle'
				self.ACTIVE[f'H{val}'].number_format = '0'
			case "Orange":
				self.ACTIVE[f'B{val}'].style = 'OrangeLightStyle'
				self.ACTIVE[f'C{val}'].style = 'PaleBlueMidStyle'
				self.ACTIVE[f'E{val}'].style = 'OrangeMidStyle'
				self.ACTIVE[f'F{val}'].style = 'OrangeMidStyle'
				self.ACTIVE[f'G{val}'].style = 'OrangeMidStyle'
				self.ACTIVE[f'H{val}'].style = 'PaleBlueMidStyle'

				self.ACTIVE[f'F{val}'].number_format = '0'
				self.ACTIVE[f'G{val}'].number_format = '0'
				self.ACTIVE[f'H{val}'].number_format = '0'
			case "Purple":
				self.ACTIVE[f'B{val}'].style = 'PurpleLightStyle'
				self.ACTIVE[f'C{val}'].style = 'PalePinkMidStyle'
				self.ACTIVE[f'E{val}'].style = 'PurpleMidStyle'
				self.ACTIVE[f'F{val}'].style = 'PurpleMidStyle'
				self.ACTIVE[f'G{val}'].style = 'PurpleMidStyle'
				self.ACTIVE[f'H{val}'].style = 'PalePinkMidStyle'

				self.ACTIVE[f'F{val}'].number_format = '0'
				self.ACTIVE[f'G{val}'].number_format = '0'
				self.ACTIVE[f'H{val}'].number_format = '0'
			case "Green":
				self.ACTIVE[f'B{val}'].style = 'GreenLightStyle'
				self.ACTIVE[f'C{val}'].style = 'BlueMidStyle'
				self.ACTIVE[f'E{val}'].style = 'GreenMidStyle'
				self.ACTIVE[f'F{val}'].style = 'GreenMidStyle'
				self.ACTIVE[f'G{val}'].style = 'GreenMidStyle'
				self.ACTIVE[f'H{val}'].style = 'BlueMidStyle'

				self.ACTIVE[f'F{val}'].number_format = '0'
				self.ACTIVE[f'G{val}'].number_format = '0'
				self.ACTIVE[f'H{val}'].number_format = '0'
			case "PaleBlue":
				self.ACTIVE[f'B{val}'].style = 'PaleBlueLightStyle'
				self.ACTIVE[f'C{val}'].style = 'GreenMidStyle'
				self.ACTIVE[f'E{val}'].style = 'PaleBlueMidStyle'
				self.ACTIVE[f'F{val}'].style = 'PaleBlueMidStyle'
				self.ACTIVE[f'G{val}'].style = 'PaleBlueMidStyle'
				self.ACTIVE[f'H{val}'].style = 'GreenMidStyle'

				self.ACTIVE[f'F{val}'].number_format = '0'
				self.ACTIVE[f'G{val}'].number_format = '0'
				self.ACTIVE[f'H{val}'].number_format = '0'
			case "PalePink":
				self.ACTIVE[f'B{val}'].style = 'PalePinkLightStyle'
				self.ACTIVE[f'C{val}'].style = 'PurpleMidStyle'
				self.ACTIVE[f'E{val}'].style = 'PalePinkMidStyle'
				self.ACTIVE[f'F{val}'].style = 'PalePinkMidStyle'
				self.ACTIVE[f'G{val}'].style = 'PalePinkMidStyle'
				self.ACTIVE[f'H{val}'].style = 'PurpleMidStyle'

				self.ACTIVE[f'F{val}'].number_format = '0'
				self.ACTIVE[f'G{val}'].number_format = '0'
				self.ACTIVE[f'H{val}'].number_format = '0'